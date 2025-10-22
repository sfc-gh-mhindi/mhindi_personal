import pandas as pd
from datetime import datetime, timedelta
import math

def calculate_moving_plan_analysis(
    starting_balance,
    settlement_date_str
):
    """
    Calculates the moving plan cash flow analysis for a single scenario.

    Args:
        starting_balance (float): The initial account balance.
        settlement_date_str (str): The settlement date (DD/MM/YYYY).

    Returns:
        tuple: (DataFrame, move_out_date, end_date, financial_summary)
    """

    # --- Input Processing and Date Conversion ---
    
    # Convert settlement date string to datetime object
    settlement_date = datetime.strptime(settlement_date_str, "%d/%m/%Y")
    
    # Calculate move out date as Saturday after the first weekend following settlement
    # First, find the end of the current week (Sunday)
    days_to_end_of_week = (6 - settlement_date.weekday()) % 7  # Sunday is weekday 6
    if days_to_end_of_week == 0:  # If settlement is on Sunday
        days_to_end_of_week = 7  # Go to next Sunday
    
    end_of_current_week = settlement_date + timedelta(days=days_to_end_of_week)
    
    # Now find the Saturday of the following week
    move_out_date = end_of_current_week + timedelta(days=6)  # Saturday of next week
    
    # Calculate end date (2 months after settlement)
    if settlement_date.month >= 11:  # November or December
        if settlement_date.month == 11:
            end_date = settlement_date.replace(year=settlement_date.year + 1, month=1)
        else:  # December
            end_date = settlement_date.replace(year=settlement_date.year + 1, month=2)
    else:
        end_date = settlement_date.replace(month=settlement_date.month + 2)
    
    # --- Cash Flow Events ---
    
    current_balance = starting_balance
    cash_flow_events = {}  # Dictionary to store events by date
    
    # Helper function to add cash flow event
    def add_cash_flow_event(date, description, amount, milestone=""):
        date_key = date.strftime("%d/%m/%Y")
        if date_key not in cash_flow_events:
            cash_flow_events[date_key] = []
        cash_flow_events[date_key].append({
            "description": description,
            "amount": amount,
            "milestone": milestone
        })
    
    # Starting balance entry
    start_date = datetime.strptime("01/08/2025", "%d/%m/%Y")
    add_cash_flow_event(start_date - timedelta(days=1), "Starting Balance", starting_balance, "INITIAL")
    
    # Income events on 01/08/2025
    income_date = start_date
    add_cash_flow_event(income_date, "Additional Savings #1", 5700, "INCOME")
    add_cash_flow_event(income_date, "Additional Savings #2", 1200, "INCOME")
    add_cash_flow_event(income_date, "Additional Savings #3", 500, "INCOME")
    
    # Settlement date expenses
    add_cash_flow_event(settlement_date, "House Settlement Payment", -52046, "SETTLEMENT")
    
    # Post-settlement expenses (relative to settlement date)
    add_cash_flow_event(settlement_date + timedelta(days=1), "Cleaning House", -200, "POST-SETTLEMENT")
    add_cash_flow_event(settlement_date + timedelta(days=1), "Fixing Windows", -300, "POST-SETTLEMENT")
    add_cash_flow_event(settlement_date + timedelta(days=2), "Bathroom Grouting & Sealing", -1000, "POST-SETTLEMENT")
    add_cash_flow_event(settlement_date + timedelta(days=2), "2-Week Rent Payment", -1400, "POST-SETTLEMENT")
    add_cash_flow_event(settlement_date + timedelta(days=3), "Ensuite Bathroom Fixes", -20000, "POST-SETTLEMENT")
    
    # Moving expenses (on the move out date - Saturday after first weekend following settlement)
    add_cash_flow_event(move_out_date, "Furniture Purchase", -3000, "MOVING")
    add_cash_flow_event(move_out_date, "Removalists", -1200, "MOVING")
    
    # --- Non-Financial Milestones ---
    
    # HYKO - Sydney events (no financial impact)
    hyko_date_1 = datetime(2025, 8, 26)
    hyko_date_2 = datetime(2025, 8, 27)
    
    add_cash_flow_event(hyko_date_1, "HYKO - Sydney (Day 1)", 0, "HYKO")
    add_cash_flow_event(hyko_date_2, "HYKO - Sydney (Day 2)", 0, "HYKO")
    
    # Additional non-financial milestones
    add_cash_flow_event(datetime(2025, 8, 15), "Building Inspection Due", 0, "INSPECTION")
    add_cash_flow_event(datetime(2025, 8, 20), "Insurance Policy Review", 0, "INSURANCE")
    add_cash_flow_event(settlement_date - timedelta(days=7), "Final Walkthrough", 0, "WALKTHROUGH")
    add_cash_flow_event(settlement_date - timedelta(days=3), "Pre-Settlement Meeting", 0, "MEETING")
    add_cash_flow_event(move_out_date - timedelta(days=1), "Packing Day", 0, "PACKING")
    add_cash_flow_event(move_out_date + timedelta(days=1), "Unpacking & Setup", 0, "UNPACKING")
    
    # --- Ongoing Expenses and Income (2 months after settlement) ---
    
    # Monthly income logic: Different amounts before and after settlement
    # Before settlement: Regular monthly savings (5700 + 500, plus 1200 for August only)
    # After settlement: Monthly income of 8261 (mortgage allocation + savings)
    
    # Handle monthly income starting from August 1st, 2025
    current_month = datetime(2025, 8, 1)  # Start from August 1st, 2025
    
    while current_month <= end_date:
        # Only add income if we haven't already added it for this month
        if current_month not in [income_date]:  # Skip if already added initial income
            if current_month < settlement_date:
                # Pre-settlement: Regular monthly savings
                add_cash_flow_event(current_month, f"Monthly Savings #1 (Pre-Settlement)", 5700, "PRE-SETTLEMENT INCOME")
                add_cash_flow_event(current_month, f"Monthly Savings #2 (Pre-Settlement)", 500, "PRE-SETTLEMENT INCOME")
                
                # Additional $1200 only for August 2025
                if current_month.month == 8 and current_month.year == 2025:
                    add_cash_flow_event(current_month, f"Monthly Savings #3 (August Only)", 1200, "PRE-SETTLEMENT INCOME")
            else:
                # Post-settlement: Full monthly income
                add_cash_flow_event(current_month, f"Monthly Income (Post-Settlement - Mortgage Allocation + Savings)", 8261, "MONTHLY INCOME")
        
        # Move to next month
        if current_month.month == 12:
            current_month = datetime(current_month.year + 1, 1, 1)
        else:
            current_month = datetime(current_month.year, current_month.month + 1, 1)
    
    # Mortgage payments: $2410 every 2 weeks starting 2 weeks after settlement
    first_mortgage_date = settlement_date + timedelta(days=14)  # 2 weeks after settlement
    mortgage_date = first_mortgage_date
    mortgage_payment_number = 1
    
    while mortgage_date <= end_date:
        add_cash_flow_event(mortgage_date, f"Mortgage Payment #{mortgage_payment_number}", -2410, "MORTGAGE")
        mortgage_date += timedelta(days=14)  # Next payment 2 weeks later
        mortgage_payment_number += 1
    
    # --- Generate Daily Cash Flow Data ---
    
    # Determine the date range for daily entries
    analysis_start_date = start_date - timedelta(days=1)  # Include starting balance date
    analysis_end_date = end_date
    
    cash_flow_data = []
    current_date = analysis_start_date
    current_balance = 0
    
    while current_date <= analysis_end_date:
        date_key = current_date.strftime("%d/%m/%Y")
        
        # Check if there are any events for this date
        if date_key in cash_flow_events:
            events = cash_flow_events[date_key]
            
            # Handle multiple events on the same date
            for i, event in enumerate(events):
                if event["description"] == "Starting Balance":
                    current_balance = event["amount"]
                else:
                    current_balance += event["amount"]
                
                # For multiple events on same date, show description for each
                if len(events) > 1:
                    description = f"{event['description']} ({i+1}/{len(events)})"
                else:
                    description = event["description"]
                
                cash_flow_data.append({
                    "Date": current_date.strftime("%d/%m/%Y"),
                    "Day of Week": current_date.strftime("%A"),
                    "Milestone": event["milestone"],
                    "Description": description,
                    "Amount": event["amount"],
                    "Running Balance": current_balance
                })
        else:
            # No events for this date, show daily balance
            cash_flow_data.append({
                "Date": current_date.strftime("%d/%m/%Y"),
                "Day of Week": current_date.strftime("%A"),
                "Milestone": "",
                "Description": "No transactions",
                "Amount": 0,
                "Running Balance": current_balance
            })
        
        current_date += timedelta(days=1)
    
    # Create DataFrame
    df = pd.DataFrame(cash_flow_data)
    
    # Store unformatted data for calculations
    df_unformatted = df.copy()
    
    # Calculate financial summary
    final_balance = df_unformatted['Running Balance'].iloc[-1]
    total_initial_income = df_unformatted[df_unformatted['Milestone'] == 'INCOME']['Amount'].sum()
    total_pre_settlement_income = df_unformatted[df_unformatted['Milestone'] == 'PRE-SETTLEMENT INCOME']['Amount'].sum()
    total_monthly_income = df_unformatted[df_unformatted['Milestone'] == 'MONTHLY INCOME']['Amount'].sum()
    total_settlement_expenses = abs(df_unformatted[df_unformatted['Milestone'] == 'SETTLEMENT']['Amount'].sum())
    total_post_settlement = abs(df_unformatted[df_unformatted['Milestone'] == 'POST-SETTLEMENT']['Amount'].sum())
    total_moving_expenses = abs(df_unformatted[df_unformatted['Milestone'] == 'MOVING']['Amount'].sum())
    total_mortgage_payments = abs(df_unformatted[df_unformatted['Milestone'] == 'MORTGAGE']['Amount'].sum())
    
    financial_summary = {
        'final_balance': final_balance,
        'total_initial_income': total_initial_income,
        'total_pre_settlement_income': total_pre_settlement_income,
        'total_monthly_income': total_monthly_income,
        'total_settlement_expenses': total_settlement_expenses,
        'total_post_settlement': total_post_settlement,
        'total_moving_expenses': total_moving_expenses,
        'total_mortgage_payments': total_mortgage_payments,
        'total_income': total_initial_income + total_pre_settlement_income + total_monthly_income,
        'total_expenses': total_settlement_expenses + total_post_settlement + total_moving_expenses + total_mortgage_payments,
        'net_change': (total_initial_income + total_pre_settlement_income + total_monthly_income) - (total_settlement_expenses + total_post_settlement + total_moving_expenses + total_mortgage_payments)
    }
    
    # Format currency columns with $ and 2 decimal places
    currency_columns = ["Amount", "Running Balance"]
    
    for col in currency_columns:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: f"${x:,.2f}" if x >= 0 else f"-${abs(x):,.2f}")
    
    # Add unformatted data as an attribute for calculations
    df.unformatted = df_unformatted
    
    return df, move_out_date, end_date, financial_summary

def generate_plan_name(settlement_date_str):
    """Generate a plan name based on settlement date."""
    settlement_date = datetime.strptime(settlement_date_str, "%d/%m/%Y")
    day_abbrev = settlement_date.strftime("%a")
    date_part = settlement_date.strftime("%d/%m")
    return f'Settlement "{day_abbrev} {date_part}"'

def create_scenario_sheet(writer, scenario_num, settlement_date_str, starting_balance, analysis_df, move_out_date, end_date, financial_summary):
    """Create a formatted sheet for a single scenario."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import calendar
    
    # Create sheet name based on date for better organization
    settlement_date = datetime.strptime(settlement_date_str, "%d/%m/%Y")
    sheet_name = f'{settlement_date.strftime("%d-%m")} Settlement'
    plan_name = generate_plan_name(settlement_date_str)
    
    # Write the analysis starting from row 80 to leave space for calendar
    analysis_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=79)
    
    # Get the worksheet
    worksheet = writer.sheets[sheet_name]
    
    # Apply month-based background colors to the data rows
    data_start_row = 80  # Data starts at row 80
    unformatted_df = analysis_df.unformatted
    
    # Define colors for different months
    month_colors = {
        7: 'E6F3FF',   # July - Light Blue
        8: 'E6FFE6',   # August - Light Green
        9: 'FFE6E6',   # September - Light Pink
        10: 'FFF2E6',  # October - Light Orange
        11: 'F0E6FF',  # November - Light Purple
        12: 'FFFACD'   # December - Light Yellow
    }
    
    current_month = None
    for idx, row in unformatted_df.iterrows():
        excel_row = data_start_row + idx + 1  # +1 because data starts after header
        date_str = row['Date']
        row_date = datetime.strptime(date_str, "%d/%m/%Y")
        row_month = row_date.month
        
        # Get color for this month
        if row_month in month_colors:
            fill_color = month_colors[row_month]
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            # Apply background color to all columns in this row
            for col in range(1, 7):  # Columns A through F
                cell = worksheet.cell(row=excel_row, column=col)
                cell.fill = fill
                
                # Add bold formatting for first entry of each month
                if current_month != row_month:
                    cell.font = Font(bold=True)
            
            current_month = row_month
    
    # Title
    worksheet['A1'] = f'SETTLEMENT {settlement_date.strftime("%d/%m/%Y")}: {plan_name.upper()}'
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet['A1'].alignment = Alignment(horizontal='center')
    worksheet.merge_cells('A1:F1')
    
    # Assumptions Section (in green)
    row = 3
    worksheet[f'A{row}'] = 'ASSUMPTIONS:'
    worksheet[f'A{row}'].font = Font(bold=True, color='008000', size=12)
    
    row += 1
    worksheet[f'A{row}'] = '• Move out date: Saturday after the first weekend following settlement'
    worksheet[f'A{row}'].font = Font(color='008000')
    
    row += 1
    worksheet[f'A{row}'] = '• Mortgage payments: $2,410 fortnightly, starting 2 weeks after settlement'
    worksheet[f'A{row}'].font = Font(color='008000')
    
    row += 1
    worksheet[f'A{row}'] = '• Pre-settlement monthly income: $6,200 ($5,700 + $500), plus $1,200 extra for August'
    worksheet[f'A{row}'].font = Font(color='008000')
    
    row += 1
    worksheet[f'A{row}'] = '• Post-settlement monthly income: $8,261 (mortgage allocation + savings)'
    worksheet[f'A{row}'].font = Font(color='008000')
    
    row += 1
    worksheet[f'A{row}'] = '• Analysis period: 2 months from settlement date'
    worksheet[f'A{row}'].font = Font(color='008000')
    
    # Plan Parameters (in black)
    row += 2
    worksheet[f'A{row}'] = 'PLAN PARAMETERS:'
    worksheet[f'A{row}'].font = Font(bold=True)
    
    row += 1
    worksheet[f'A{row}'] = f'Starting Balance:'
    worksheet[f'B{row}'] = f'${starting_balance:,.2f}'
    
    row += 1
    worksheet[f'A{row}'] = f'Settlement Date:'
    worksheet[f'B{row}'] = f'{settlement_date_str} ({settlement_date.strftime("%A")})'
    
    row += 1
    worksheet[f'A{row}'] = f'Move Out Date:'
    worksheet[f'B{row}'] = f'{move_out_date.strftime("%d/%m/%Y")} ({move_out_date.strftime("%A")})'
    
    row += 1
    worksheet[f'A{row}'] = f'Analysis Period:'
    worksheet[f'B{row}'] = f'Until {end_date.strftime("%d/%m/%Y")} (2 months after settlement)'
    
    # Financial Summary (in blue)
    row += 2
    worksheet[f'A{row}'] = 'FINANCIAL SUMMARY:'
    worksheet[f'A{row}'].font = Font(bold=True, color='0000FF', size=12)
    
    row += 1
    worksheet[f'A{row}'] = f'Initial Additional Income:'
    worksheet[f'A{row}'].font = Font(color='0000FF')
    worksheet[f'B{row}'] = f'${financial_summary["total_initial_income"]:,.2f}'
    worksheet[f'B{row}'].font = Font(color='0000FF')
    
    row += 1
    worksheet[f'A{row}'] = f'Pre-Settlement Monthly Income:'
    worksheet[f'A{row}'].font = Font(color='0000FF')
    worksheet[f'B{row}'] = f'${financial_summary["total_pre_settlement_income"]:,.2f}'
    worksheet[f'B{row}'].font = Font(color='0000FF')
    
    row += 1
    worksheet[f'A{row}'] = f'Post-Settlement Monthly Income:'
    worksheet[f'A{row}'].font = Font(color='0000FF')
    worksheet[f'B{row}'] = f'${financial_summary["total_monthly_income"]:,.2f}'
    worksheet[f'B{row}'].font = Font(color='0000FF')
    
    row += 1
    worksheet[f'A{row}'] = f'Settlement Expenses:'
    worksheet[f'A{row}'].font = Font(color='0000FF')
    worksheet[f'B{row}'] = f'${financial_summary["total_settlement_expenses"]:,.2f}'
    worksheet[f'B{row}'].font = Font(color='0000FF')
    
    row += 1
    worksheet[f'A{row}'] = f'Moving & Setup Expenses:'
    worksheet[f'A{row}'].font = Font(color='0000FF')
    worksheet[f'B{row}'] = f'${financial_summary["total_post_settlement"] + financial_summary["total_moving_expenses"]:,.2f}'
    worksheet[f'B{row}'].font = Font(color='0000FF')
    
    row += 1
    worksheet[f'A{row}'] = f'Mortgage Payments:'
    worksheet[f'A{row}'].font = Font(color='0000FF')
    worksheet[f'B{row}'] = f'${financial_summary["total_mortgage_payments"]:,.2f}'
    worksheet[f'B{row}'].font = Font(color='0000FF')
    
    row += 1
    worksheet[f'A{row}'] = f'Net Change:'
    worksheet[f'A{row}'].font = Font(color='0000FF', bold=True)
    worksheet[f'B{row}'] = f'${financial_summary["net_change"]:,.2f}'
    worksheet[f'B{row}'].font = Font(color='0000FF', bold=True)
    
    # Final Result (in red)
    row += 2
    worksheet[f'A{row}'] = 'FINAL RESULT:'
    worksheet[f'A{row}'].font = Font(bold=True, color='FF0000', size=12)
    
    row += 1
    worksheet[f'A{row}'] = f'Final Balance:'
    worksheet[f'A{row}'].font = Font(color='FF0000', bold=True)
    worksheet[f'B{row}'] = f'${financial_summary["final_balance"]:,.2f}'
    worksheet[f'B{row}'].font = Font(color='FF0000', bold=True, size=11)
    
    # Status check
    row += 1
    if financial_summary["final_balance"] >= 0:
        status = "✓ PLAN VIABLE - Sufficient funds available"
        status_color = '008000'  # Green
    else:
        status = "⚠ PLAN REQUIRES ATTENTION - Insufficient funds"
        status_color = 'FF0000'  # Red
    
    worksheet[f'A{row}'] = f'Status:'
    worksheet[f'A{row}'].font = Font(color=status_color, bold=True)
    worksheet[f'B{row}'] = status
    worksheet[f'B{row}'].font = Font(color=status_color, bold=True)
    worksheet.merge_cells(f'B{row}:F{row}')
    
    # --- CREATE CALENDAR VIEW ---
    
    # Create calendar data structure
    calendar_data = {}
    
    # Process the unformatted data to extract calendar information
    for idx, row_data in unformatted_df.iterrows():
        date_str = row_data['Date']
        row_date = datetime.strptime(date_str, "%d/%m/%Y")
        
        # Create month key
        month_key = (row_date.year, row_date.month)
        
        if month_key not in calendar_data:
            calendar_data[month_key] = {}
        
        day = row_date.day
        
        # Initialize day data if not exists
        if day not in calendar_data[month_key]:
            calendar_data[month_key][day] = {
                'balance': row_data['Running Balance'],
                'movement': row_data['Amount'],
                'milestones': [],
                'descriptions': []
            }
        
        # Add milestone and description if not "No transactions"
        if row_data['Description'] != "No transactions":
            if row_data['Milestone']:
                calendar_data[month_key][day]['milestones'].append(row_data['Milestone'])
            calendar_data[month_key][day]['descriptions'].append(row_data['Description'])
        
        # Update balance (keep the latest balance for the day)
        calendar_data[month_key][day]['balance'] = row_data['Running Balance']
    
    # Create calendar views
    calendar_start_row = row + 3
    
    # Calendar header
    worksheet[f'A{calendar_start_row}'] = 'MONTHLY CALENDAR VIEW:'
    worksheet[f'A{calendar_start_row}'].font = Font(bold=True, size=12)
    
    calendar_row = calendar_start_row + 2
    
    # Create calendars for each month in the analysis period
    for month_key in sorted(calendar_data.keys()):
        year, month = month_key
        month_name = calendar.month_name[month]
        
        # Month header
        worksheet[f'A{calendar_row}'] = f'{month_name} {year}'
        worksheet[f'A{calendar_row}'].font = Font(bold=True, size=11, color='000080')
        worksheet.merge_cells(f'A{calendar_row}:G{calendar_row}')
        calendar_row += 1
        
        # Calendar legend
        worksheet[f'A{calendar_row}'] = 'Legend: Balance | Movement | Milestone'
        worksheet[f'A{calendar_row}'].font = Font(size=9, italic=True)
        worksheet.merge_cells(f'A{calendar_row}:G{calendar_row}')
        calendar_row += 1
        
        # Day headers
        day_headers = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        for col, day_header in enumerate(day_headers, 1):
            cell = worksheet.cell(row=calendar_row, column=col)
            cell.value = day_header
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        calendar_row += 1
        
        # Get calendar for the month
        cal = calendar.monthcalendar(year, month)
        
        # Create calendar grid
        for week in cal:
            for col, day in enumerate(week, 1):
                cell = worksheet.cell(row=calendar_row, column=col)
                
                # Set cell dimensions
                worksheet.column_dimensions[get_column_letter(col)].width = 15
                worksheet.row_dimensions[calendar_row].height = 60
                
                if day == 0:
                    # Empty cell for days not in this month
                    cell.value = ""
                    cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                else:
                    # Get data for this day
                    day_data = calendar_data[month_key].get(day, {
                        'balance': 0,
                        'movement': 0,
                        'milestones': [],
                        'descriptions': []
                    })
                    
                    # Format cell content
                    balance = day_data['balance']
                    movement = day_data['movement']
                    milestones = day_data['milestones']
                    
                    # Create cell content
                    cell_content = f"{day}\n"
                    cell_content += f"${balance:,.0f}\n"
                    
                    if movement != 0:
                        if movement > 0:
                            cell_content += f"+${movement:,.0f}\n"
                        else:
                            cell_content += f"-${abs(movement):,.0f}\n"
                    else:
                        cell_content += "No change\n"
                    
                    # Add milestone abbreviations
                    milestone_abbrev = {
                        'INITIAL': 'INIT',
                        'INCOME': 'INC',
                        'SETTLEMENT': 'SETT',
                        'POST-SETTLEMENT': 'POST',
                        'MOVING': 'MOVE',
                        'PRE-SETTLEMENT INCOME': 'PRE-INC',
                        'MONTHLY INCOME': 'M-INC',
                        'MORTGAGE': 'MORT',
                        'HYKO': 'HYKO',
                        'INSPECTION': 'INS',
                        'INSURANCE': 'INS-POL',
                        'WALKTHROUGH': 'WALK',
                        'MEETING': 'MEET',
                        'PACKING': 'PACK',
                        'UNPACKING': 'UNPACK'
                    }
                    
                    if milestones:
                        unique_milestones = list(set(milestones))
                        milestone_text = ', '.join([milestone_abbrev.get(m, m[:4]) for m in unique_milestones])
                        cell_content += milestone_text
                    
                    cell.value = cell_content
                    cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                    cell.font = Font(size=11)
                    
                    # Color coding based on financial impact
                    if any(m in ['HYKO', 'INSPECTION', 'INSURANCE', 'WALKTHROUGH', 'MEETING', 'PACKING', 'UNPACKING'] for m in milestones):
                        # Yellow for non-financial milestones (check this first)
                        cell.fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
                    elif movement > 0:
                        # Green for positive financial impact (increases)
                        cell.fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')
                    elif movement < 0:
                        # Red for negative financial impact (decreases)
                        cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                    else:
                        # White for no transactions
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                # Add borders to all cells
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
            
            calendar_row += 1
        
        # Add space between months
        calendar_row += 2
    
    # Calendar color legend
    worksheet[f'A{calendar_row}'] = 'Calendar Color Legend:'
    worksheet[f'A{calendar_row}'].font = Font(bold=True, size=10)
    calendar_row += 1
    
    legend_items = [
        ('Positive Financial Impact (+)', 'E6FFE6'),
        ('Negative Financial Impact (-)', 'FFE6E6'),
        ('Non-Financial Milestones', 'FFFACD'),
        ('No Transactions', 'FFFFFF')
    ]
    
    for item, color in legend_items:
        cell = worksheet.cell(row=calendar_row, column=1)
        cell.value = item
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.font = Font(size=9)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        calendar_row += 1
    
    # Add a clear separator line before the detailed table
    calendar_row += 1
    worksheet[f'A{calendar_row}'] = '=' * 80
    worksheet[f'A{calendar_row}'].font = Font(bold=True)
    worksheet.merge_cells(f'A{calendar_row}:F{calendar_row}')
    
    # Add table header
    calendar_row += 1
    worksheet[f'A{calendar_row}'] = 'DETAILED DAILY CASH FLOW ANALYSIS:'
    worksheet[f'A{calendar_row}'].font = Font(bold=True, size=11)
    
    # Add color legend for detailed table
    calendar_row += 1
    worksheet[f'A{calendar_row}'] = 'Color Legend:'
    worksheet[f'A{calendar_row}'].font = Font(bold=True, size=10)
    
    # Create legend with actual colors
    legend_items = [
        ('July', 'E6F3FF'),
        ('August', 'E6FFE6'),
        ('September', 'FFE6E6'),
        ('October', 'FFF2E6'),
        ('November', 'F0E6FF'),
        ('December', 'FFFACD')
    ]
    
    col_offset = 0
    for month_name, color_code in legend_items:
        legend_cell = worksheet.cell(row=calendar_row, column=2 + col_offset)
        legend_cell.value = month_name
        legend_cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
        legend_cell.font = Font(size=9, bold=True)
        legend_cell.alignment = Alignment(horizontal='center')
        col_offset += 1
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        if column_letter and max_length > 0:
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

def create_summary_sheet(writer, scenarios_data, starting_balance):
    """Create a summary comparison sheet for all scenarios."""
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Create summary data
    summary_data = []
    for i, (settlement_date, _, _, _, financial_summary) in enumerate(scenarios_data, 1):
        plan_name = generate_plan_name(settlement_date)
        settlement_dt = datetime.strptime(settlement_date, "%d/%m/%Y")
        
        summary_data.append({
            'Scenario': f'Scenario {i}',
            'Settlement Date': settlement_date,
            'Day of Week': settlement_dt.strftime("%A"),
            'Plan Name': plan_name,
            'Final Balance': financial_summary['final_balance'],
            'Total Income': financial_summary['total_income'],
            'Total Expenses': financial_summary['total_expenses'],
            'Net Change': financial_summary['net_change'],
            'Status': 'VIABLE' if financial_summary['final_balance'] >= 0 else 'REQUIRES ATTENTION'
        })
    
    # Create DataFrame
    summary_df = pd.DataFrame(summary_data)
    
    # Format currency columns
    currency_columns = ['Final Balance', 'Total Income', 'Total Expenses', 'Net Change']
    for col in currency_columns:
        summary_df[col] = summary_df[col].apply(lambda x: f"${x:,.2f}" if x >= 0 else f"-${abs(x):,.2f}")
    
    # Write to Excel
    summary_df.to_excel(writer, sheet_name='Summary Comparison', index=False, startrow=10)
    
    # Get the worksheet
    worksheet = writer.sheets['Summary Comparison']
    
    # Apply month-based background colors to summary data
    month_colors = {
        7: 'E6F3FF',   # July - Light Blue
        8: 'E6FFE6',   # August - Light Green
        9: 'FFE6E6',   # September - Light Pink
        10: 'FFF2E6',  # October - Light Orange
        11: 'F0E6FF',  # November - Light Purple
        12: 'FFFACD'   # December - Light Yellow
    }
    
    # Apply colors to summary rows based on settlement month
    for idx, (settlement_date, _, _, _, _) in enumerate(scenarios_data):
        excel_row = 11 + idx + 1  # Data starts at row 11, +1 for header
        settlement_dt = datetime.strptime(settlement_date, "%d/%m/%Y")
        settlement_month = settlement_dt.month
        
        if settlement_month in month_colors:
            fill_color = month_colors[settlement_month]
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            # Apply background color to all columns in this row
            for col in range(1, 10):  # Columns A through I
                cell = worksheet.cell(row=excel_row, column=col)
                cell.fill = fill
    
    # Title
    worksheet['A1'] = 'MOVING PLAN SCENARIO ANALYSIS - SUMMARY COMPARISON'
    worksheet['A1'].font = Font(bold=True, size=16)
    worksheet['A1'].alignment = Alignment(horizontal='center')
    worksheet.merge_cells('A1:I1')
    
    # Parameters
    row = 3
    worksheet[f'A{row}'] = 'ANALYSIS PARAMETERS:'
    worksheet[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    worksheet[f'A{row}'] = f'Starting Balance: ${starting_balance:,.2f}'
    worksheet[f'A{row}'].font = Font(size=11)
    
    row += 1
    worksheet[f'A{row}'] = f'Analysis Period: 2 months from each settlement date'
    worksheet[f'A{row}'].font = Font(size=11)
    
    row += 1
    worksheet[f'A{row}'] = f'Monthly Income: $8,261 (mortgage allocation + savings)'
    worksheet[f'A{row}'].font = Font(size=11)
    
    row += 1
    worksheet[f'A{row}'] = f'Mortgage Payments: $2,410 fortnightly'
    worksheet[f'A{row}'].font = Font(size=11)
    
    # Table header
    row += 2
    worksheet[f'A{row}'] = 'SCENARIO COMPARISON:'
    worksheet[f'A{row}'].font = Font(bold=True, size=12)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        if column_letter and max_length > 0:
            adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

def create_chart_sheet(writer, scenarios_data, starting_balance):
    """Create a chart worksheet showing balance by date for all scenarios."""
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.marker import DataPoint
    import pandas as pd
    from openpyxl.chart.data_source import NumData, NumVal
    from openpyxl.chart.label import DataLabelList
    
    # Prepare data for charting
    all_dates = set()
    scenario_data = {}
    
    # Collect all unique dates and scenario data
    for i, (settlement_date, analysis_df, _, _, _) in enumerate(scenarios_data, 1):
        scenario_name = f"Scenario {i} ({settlement_date})"
        unformatted_df = analysis_df.unformatted
        
        # Convert dates to datetime for proper sorting
        unformatted_df['Date_dt'] = pd.to_datetime(unformatted_df['Date'], format='%d/%m/%Y')
        unformatted_df = unformatted_df.sort_values('Date_dt')
        
        scenario_data[scenario_name] = unformatted_df
        all_dates.update(unformatted_df['Date_dt'].tolist())
    
    # Create a comprehensive date range
    all_dates = sorted(list(all_dates))
    
    # Create chart data structure
    chart_data = []
    
    for date in all_dates:
        row = {'Date': date.strftime('%d/%m/%Y')}
        
        for scenario_name, df in scenario_data.items():
            # Find the balance for this date (or carry forward the last known balance)
            date_matches = df[df['Date_dt'] <= date]
            if not date_matches.empty:
                balance = date_matches.iloc[-1]['Running Balance']
                row[scenario_name] = balance
            else:
                row[scenario_name] = starting_balance
        
        chart_data.append(row)
    
    # Create DataFrame for chart
    chart_df = pd.DataFrame(chart_data)
    
    # Create milestone data for annotations
    milestone_data = []
    for i, (settlement_date, analysis_df, move_out_date, end_date, financial_summary) in enumerate(scenarios_data, 1):
        scenario_name = f"Scenario {i} ({settlement_date})"
        unformatted_df = analysis_df.unformatted
        
        # Key milestones
        settlement_dt = datetime.strptime(settlement_date, "%d/%m/%Y")
        milestones = [
            (settlement_dt, "Settlement", financial_summary['final_balance']),
            (move_out_date, "Move Out", None),
        ]
        
        # Add mortgage payment dates
        mortgage_dates = unformatted_df[unformatted_df['Milestone'] == 'MORTGAGE']
        for _, row in mortgage_dates.iterrows():
            milestone_dt = datetime.strptime(row['Date'], "%d/%m/%Y")
            milestones.append((milestone_dt, "Mortgage", row['Running Balance']))
        
        # Add monthly income dates
        income_dates = unformatted_df[unformatted_df['Milestone'] == 'MONTHLY INCOME']
        for _, row in income_dates.iterrows():
            milestone_dt = datetime.strptime(row['Date'], "%d/%m/%Y")
            milestones.append((milestone_dt, "Income", row['Running Balance']))
        
        # Add HYKO dates
        hyko_date_1 = datetime(2025, 8, 26)
        hyko_date_2 = datetime(2025, 8, 27)
        milestones.append((hyko_date_1, "HYKO", None))
        milestones.append((hyko_date_2, "HYKO", None))
        
        milestone_data.append((scenario_name, milestones))
    
    # Write chart data to Excel
    chart_df.to_excel(writer, sheet_name='Balance Chart', index=False, startrow=10)
    
    # Get the worksheet
    worksheet = writer.sheets['Balance Chart']
    
    # Title
    worksheet['A1'] = 'MOVING PLAN SCENARIO ANALYSIS - BALANCE BY DATE CHART'
    worksheet['A1'].font = Font(bold=True, size=16)
    worksheet['A1'].alignment = Alignment(horizontal='center')
    worksheet.merge_cells('A1:G1')
    
    # Parameters
    row = 3
    worksheet[f'A{row}'] = 'CHART OVERVIEW:'
    worksheet[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    worksheet[f'A{row}'] = f'Starting Balance: ${starting_balance:,.2f}'
    worksheet[f'A{row}'].font = Font(size=11)
    
    row += 1
    worksheet[f'A{row}'] = f'Analysis Period: 2 months from each settlement date'
    worksheet[f'A{row}'].font = Font(size=11)
    
    row += 1
    worksheet[f'A{row}'] = 'Key Milestones: Settlement (red), Move Out (blue), Mortgage Payments (orange), Monthly Income (green)'
    worksheet[f'A{row}'].font = Font(size=11)
    
    # Table header
    row += 2
    worksheet[f'A{row}'] = 'BALANCE DATA BY DATE:'
    worksheet[f'A{row}'].font = Font(bold=True, size=12)
    
    # Create line chart
    chart = LineChart()
    chart.title = "Account Balance Over Time - All Scenarios"
    chart.style = 10
    chart.x_axis.title = 'Date'
    chart.y_axis.title = 'Account Balance ($)'
    chart.width = 20
    chart.height = 12
    
    # Add data to chart
    data_start_row = 11  # Data starts at row 11
    data_end_row = data_start_row + len(chart_df) - 1
    
    # Add each scenario as a series
    for col_idx, col_name in enumerate(chart_df.columns[1:], 2):  # Skip Date column
        data_ref = Reference(worksheet, min_col=col_idx, min_row=data_start_row, max_col=col_idx, max_row=data_end_row)
        series = chart.add_data(data_ref, titles_from_data=True)
    
    # Set category axis (dates)
    cats = Reference(worksheet, min_col=1, min_row=data_start_row+1, max_row=data_end_row)
    chart.set_categories(cats)
    
    # Configure chart appearance
    chart.legend.position = 'r'  # Right side legend
    chart.x_axis.tickLblPos = 'low'
    chart.y_axis.tickLblPos = 'low'
    
    # Format y-axis to show currency
    chart.y_axis.numFmt = '$#,##0'
    
    # Add grid lines properly
    from openpyxl.chart.axis import ChartLines
    chart.y_axis.majorGridlines = ChartLines()  # Add horizontal grid lines
    chart.x_axis.majorGridlines = ChartLines()  # Add vertical grid lines
    
    # Add data labels and markers to each series
    for i, series in enumerate(chart.series):
        # Create data labels
        labels = DataLabelList()
        labels.showVal = True
        labels.showSerName = False
        labels.showCatName = False
        labels.showPercent = False
        labels.numFmt = '$#,##0'
        labels.position = 't'  # 't' for top position
        
        # Only show labels for smaller datasets to avoid clutter
        if len(chart_df) <= 15:
            series.dLbls = labels
        
        # Add markers to make lines more visible
        series.marker.symbol = 'circle'
        series.marker.size = 5
        series.smooth = True  # Smooth lines for better appearance
    
    # Position chart
    worksheet.add_chart(chart, "A25")
    
    # Add milestone annotations below the chart
    milestone_row = 50
    worksheet[f'A{milestone_row}'] = 'KEY MILESTONES BY SCENARIO:'
    worksheet[f'A{milestone_row}'].font = Font(bold=True, size=12)
    
    milestone_row += 2
    for scenario_name, milestones in milestone_data:
        worksheet[f'A{milestone_row}'] = scenario_name
        worksheet[f'A{milestone_row}'].font = Font(bold=True, size=11)
        milestone_row += 1
        
        # Sort milestones by date
        sorted_milestones = sorted(milestones, key=lambda x: x[0])
        
        for milestone_date, milestone_type, balance in sorted_milestones:
            if balance is not None:
                worksheet[f'B{milestone_row}'] = f"{milestone_date.strftime('%d/%m/%Y')}: {milestone_type} - ${balance:,.2f}"
            else:
                worksheet[f'B{milestone_row}'] = f"{milestone_date.strftime('%d/%m/%Y')}: {milestone_type}"
            
            # Color code milestones
            if milestone_type == "Settlement":
                worksheet[f'B{milestone_row}'].font = Font(color='FF0000')  # Red
            elif milestone_type == "Move Out":
                worksheet[f'B{milestone_row}'].font = Font(color='0000FF')  # Blue
            elif milestone_type == "Mortgage":
                worksheet[f'B{milestone_row}'].font = Font(color='FF8000')  # Orange
            elif milestone_type == "Income":
                worksheet[f'B{milestone_row}'].font = Font(color='008000')  # Green
            elif milestone_type == "HYKO":
                worksheet[f'B{milestone_row}'].font = Font(color='0000FF') # Blue for HYKO
            
            milestone_row += 1
        
        milestone_row += 1  # Extra space between scenarios
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        if column_letter and max_length > 0:
            adjusted_width = min(max_length + 2, 40)  # Cap at 40 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

if __name__ == "__main__":
    print("--- Moving Plan Scenario Analysis ---")

    # Get inputs from the user with default values
    try:
        starting_balance_input = input("Enter Starting Balance [default: $93,000]: $").strip()
        starting_balance = float(starting_balance_input.replace(',', '')) if starting_balance_input else 93000
        
        print("\nEnter 5 Settlement Dates for Scenario Analysis:")
        print("Default dates: 01/08/2025, 08/08/2025, 15/08/2025, 22/08/2025, 02/09/2025")
        
        default_dates = ["01/08/2025", "08/08/2025", "15/08/2025", "22/08/2025", "02/09/2025"]
        settlement_dates = []
        
        for i in range(5):
            date_input = input(f"Enter Settlement Date {i+1} (DD/MM/YYYY) [default: {default_dates[i]}]: ").strip()
            settlement_date = date_input if date_input else default_dates[i]
            
            # Validate settlement date format
            try:
                datetime.strptime(settlement_date, "%d/%m/%Y")
                settlement_dates.append(settlement_date)
            except ValueError:
                print(f"Invalid date format for date {i+1}. Using default: {default_dates[i]}")
                settlement_dates.append(default_dates[i])

    except ValueError:
        print("Invalid input. Please ensure all numerical inputs are valid numbers.")
        exit()

    # Calculate scenarios
    try:
        print("\nCalculating scenarios...")
        scenarios_data = []
        
        for i, settlement_date in enumerate(settlement_dates, 1):
            print(f"Processing Scenario {i}: {settlement_date}")
            analysis_df, move_out_date, end_date, financial_summary = calculate_moving_plan_analysis(
                starting_balance, settlement_date
            )
            scenarios_data.append((settlement_date, analysis_df, move_out_date, end_date, financial_summary))
        
        # Sort scenarios by date for ordered Excel sheets
        scenarios_data_sorted = sorted(scenarios_data, key=lambda x: datetime.strptime(x[0], "%d/%m/%Y"))
        
        # Save to Excel with multiple sheets
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"moving_plan_scenario_analysis_{timestamp}.xlsx"
        
        # Create Excel writer object for custom formatting
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # Create individual scenario sheets first (ordered by date)
            for i, (settlement_date, analysis_df, move_out_date, end_date, financial_summary) in enumerate(scenarios_data_sorted, 1):
                create_scenario_sheet(writer, i, settlement_date, starting_balance, analysis_df, move_out_date, end_date, financial_summary)
            
            # Create summary sheet
            create_summary_sheet(writer, scenarios_data, starting_balance)
            
            # Create chart sheet as the final worksheet
            create_chart_sheet(writer, scenarios_data, starting_balance)

        print(f"\nScenario analysis generated successfully! Saved to '{output_filename}'")
        print("\nSCENARIO SUMMARY:")
        print("=" * 80)
        
        for i, (settlement_date, _, move_out_date, end_date, financial_summary) in enumerate(scenarios_data, 1):
            plan_name = generate_plan_name(settlement_date)
            final_balance = financial_summary['final_balance']
            status = "✓ VIABLE" if final_balance >= 0 else "⚠ REQUIRES ATTENTION"
            
            print(f"Scenario {i}: {plan_name}")
            print(f"  Settlement Date: {settlement_date}")
            print(f"  Move Out Date: {move_out_date.strftime('%d/%m/%Y')} ({move_out_date.strftime('%A')})")
            print(f"  Final Balance: ${final_balance:,.2f}")
            print(f"  Status: {status}")
            print()
        
        # Find best and worst scenarios
        best_scenario = max(scenarios_data, key=lambda x: x[4]['final_balance'])
        worst_scenario = min(scenarios_data, key=lambda x: x[4]['final_balance'])
        
        best_idx = scenarios_data.index(best_scenario) + 1
        worst_idx = scenarios_data.index(worst_scenario) + 1
        
        print("RECOMMENDATIONS:")
        print(f"🏆 Best Scenario: Scenario {best_idx} ({best_scenario[0]}) - ${best_scenario[4]['final_balance']:,.2f}")
        print(f"⚠️  Worst Scenario: Scenario {worst_idx} ({worst_scenario[0]}) - ${worst_scenario[4]['final_balance']:,.2f}")
        
        if all(scenario[4]['final_balance'] >= 0 for scenario in scenarios_data):
            print("✅ All scenarios are financially viable!")
        else:
            viable_count = sum(1 for scenario in scenarios_data if scenario[4]['final_balance'] >= 0)
            print(f"📊 {viable_count} out of 5 scenarios are viable.")
        
    except Exception as e:
        print(f"\nAn error occurred: {e}")
        import traceback
        traceback.print_exc() 